import numpy as np
import pandas as pd
from openpyxl import load_workbook
import pickle


class DataHandler:
    """
    Extract data from an excel file containing student activities.

    This class opens a file containing the data from multiple students on how
    they performed on the Snappet app. It extracts the user ID's, learning
    objective ID's, exercise ID's, date and time and correctness of a
    students answer on an exercise. It tracks what kind of exercise each
    particular exercise is.

    Additionally an instance of the moment by moment calculator class is
    created.
    """
    def __init__(self,
                 fname="Data Simone.xlsx",
                 sheetname="Onlyfirstattempts"):
        # Load the workbook.
        print("loading workbook")
        self.load_student_workbook(fname)

        # Set pre calculated brute force parameters
        self.l0s = [0.852, 0.256, 0.091, .690]  # From bruteforceparameters
        self.ts = [0.033, 0.318, 0.158, .205]
        self.gs = [0.299, 0.043, 0.116, .112]
        self.ss = [0.099, 0.099, 0.099, .099]
        self.fs = [0.042, 0.064, 0.028, 0.012]

        self.ol0s = [.397,   .642,   .919]      # For simone data
        self.ots = [.095,   .144,   .184]
        self.ogs = [.3,     .117,   .001]
        self.oss = [.1,     .1,     .077]

        # self.ol0s = [1, 0.001, .027, 0.536]  # From bruteforceparameters
        # self.ots = [1, 0.149, .059, 0.101]
        # self.ogs = [.3, 0.299, 0.250, 0.232]
        # self.oss = [0.1, 0.1, 0.1, 0.1]

        # Retrieve the color belonging to the exercise IDS.
        self.pre_ids, self.c_in_ids, self.c_ex_ids, self.a_ex_ids, \
            self.ra_ex_ids, self.post_ids = self.get_color_ids()

        # Initialize moment by moment class instance.
        self.m2m = MomentByMoment(self.user_ids, self.corrects, self)

    def load_student_workbook(self, fname, sheetname=None):
        """
        Loads the workbook with the student actions, if possible from pickle.
        Note: The make up of the workbook should fit in what columns are being
        extracted.

        :param fname: the path to the excel file.
        """
        if sheetname is None:
            excel_file = fname
            while fname[-1] != '.':
                fname = fname[:-1]
            fname = fname.split('/')[-1]+"pkl"
            self.max_row = 8088  # self.get_max_row()
            try:
                print("Trying to open {}".format(fname))
                self.dates, self.times, self.user_ids, self.learn_obj_ids, \
                self.exercise_ids, self.corrects, self.ability_scores \
                    = pickle.load(open(fname, "rb"))
            except FileNotFoundError:
                print("Failed opening {}\nInstead opening {}".format(fname,
                                                                     excel_file))
                wb = load_workbook(excel_file)
                self.ws = wb.active
                self.dates = self.get_column(2)
                self.times = self.get_column(3)
                self.user_ids = self.get_column(5)
                self.learn_obj_ids = self.get_column(7)
                self.exercise_ids = self.get_column(6)
                self.corrects = self.get_column(8)
                self.ability_scores = self.get_column(9)
                pickle.dump([self.dates, self.times, self.user_ids,
                             self.learn_obj_ids, self.exercise_ids,
                             self.corrects, self.ability_scores], open(fname,
                                                                       "wb"))
        else:
            return pd.read_excel(fname, sheetname)

    def get_max_row(self, column=1):
        """
        Gets the maximum row value in the excel file.

        Needed to bound the get_column function
        :param column: integer representing for what column we find the last
            row.
        :return: integer representing the last column with a value.
        """
        val = self.ws.cell(row=1, column=column).value
        i = 1
        while val:
            i += 1
            val = self.ws.cell(row=i, column=column).value
        return i

    def get_color_row(self, column=1):
        """
        Get's the index of the row where the color is located.

        Needed to split the indices for the exercises in different colors.

        :param column: The column in which we want to find the change.
        :return: integer representing the row index where the color changes.
        """
        for i in range(1, 20):
            # Find the color of the cell.
            color = self.ws.cell(row=i, column=column).fill.start_color.index
            if color != '00000000':
                return i

    def get_column(self, cid, start_row=8, end_row=None):
        """
        Get the values of a certain column.

        :param cid: integer; the ID of the column from which we
            get the variables
        :param start_row: integer; the ID of which row we start.
        :param end_row: integer; the ID of what the last row is that we
            include. defaults to self.max_row
        :return: numpy array with the values in the specified column.
        """
        # Set default value
        if end_row is None:
            end_row = self.max_row

        # Check whether parameters are valid, else return empty array
        if end_row <= start_row:
            return np.array([])

        # Retrieve the values in the column
        column = []
        for i in range(start_row, end_row):
            # Get value from cell.
            value = self.ws.cell(row=i, column=cid).value
            # Make sure the value is implemented correctly.
            if value or value == 0:
                column.append(value)
            if value is None:
                column.append(None)
        return np.array(column)

    def get_users(self):
        """
        Get the unique ID's of all users.
        :return: numpy array containing all users once.
        """
        return np.unique(self.user_ids)

    # TODO: check whether still used. if so implement more methods.
    # def get_corrects_from_user(self, user_id=0, method='all'):
    #     """
    #     Get the answers from a user in the column correct.
    #
    #     :param user_id: string representing the user.
    #     :param method: what type of answers should be given.
    #     :return: the answers given by a user according to a certain method.
    #     """
    #     if method == 'all':
    #         return self.corrects[np.where(self.user_ids == user_id)]

    def get_graph_variables(self, user_id=0, method='all', oid=None):
        """
        Generates the variables that will be shown in the graph.

        Create the boundaries and colors for the list representing the phase of
        the lesson. get the P(J) values per answer of a student and return that
        as a list.
        Method handles what happens when a student tries multiple answers
        per exercise. The options are:
        - all: use all answers
        - first: Use only the first answer on any exercise
        - second: Use only the second answer on any exercise
        - all but first: Use all but the first answer on any exercise.
        - last: use the last answer on any exercise.
        :param user_id: String representing the student.
        :param method: String representing the method of selection of answers.
        :param oid:
        :return:
        """
        print("getting variables for user {}".format(user_id))
        if oid is not None:
            print("and for skill id {}".format(oid))
        self.set_ps_correct(oid)
        p_jn, p_jl, p_jf, o_p_j, split, answers = self.m2m.get_p_js(
            user_id=user_id, method=method, objective_id=oid)
        self.graph_length = len(p_jl)
        self.boundary_list, self.color_list = self.m2m.get_color_bars()
        return [p_jn, p_jl, p_jf, o_p_j, split, answers]

    def set_ps_correct_calc(self, oid):
        """ Method to generate the pre calculated parameters. """
        loids = np.where(self.learn_obj_ids == oid)
        sames = [0]
        answers = [self.corrects[loids[0][0]]]
        for l, nextl in zip(loids[0][:-1], loids[0][1:]):
            if self.user_ids[l] == self.user_ids[nextl]:
                sames.append(1)
            else:
                sames.append(0)
            answers.append(self.corrects[nextl])
        f = 0
        ol, ot, og, os = ParameterExtractor().smart_ssr(answers, sames,
                                                      1000, 10) #old params
        l, t, g, s, f = [0, 0, 0, 0, 0]#\
            # ParameterExtractor_with_forget().smart_ssr_f(answers,
            #                                                          sames,
            #                                                          1000,
            #                                                          10)
        self.m2m.set_ps(l, t, g, s, f, ol, ot, og, os)

    def set_ps_correct(self, oid):
        """
        Set the corresponding precalculated parameters for the learning goal.
        :param oid: string representing which learning goal is used.
        """
        loids = ['7771', '7789', '8025', '7579']  # hardcoded for this file
        position = loids.index(str(oid).replace("a", ""))

        l = self.l0s[position]
        t = self.ts[position]
        g = self.gs[position]
        s = self.ss[position]
        f = self.fs[position]

        ol = self.ol0s[position]
        ot = self.ots[position]
        og = self.ogs[position]
        os = self.oss[position]
        self.m2m.set_ps(l, t, g, s, f, ol, ot, og, os)

    def get_color_ids(self, fname='res/ID exercises.xlsx'):
        """
        Set what exercises belong to what phase in the lesson.
        :param fname: Path to the file containing information about the
            exercises.
        :return: per lesson phase a list of the corresponding ID's
        """
        wb = load_workbook(fname)
        self.ws = wb.active
        pre = list(self.get_column(1, 3))
        c_bound = [self.get_color_row(i) for i in range(2, 5)]
        cin = list(self.get_column(2, 3, c_bound[0])) + \
              list(self.get_column(3, 3, c_bound[0])) + \
              list(self.get_column(4, 3, c_bound[0]))
        cex = list(self.get_column(2, c_bound[0])) + \
              list(self.get_column(3, c_bound[0])) + \
              list(self.get_column(4, c_bound[0]))
        aex = []
        raex = []
        post = list(self.get_column(5, 3))
        return (pre, cin, cex, aex, raex, post)


class MomentByMoment:
    def __init__(self, user_ids, corrects, handler: DataHandler):
        self.p_l0 = 0.064
        self.p_T = 0.095
        self.p_G = 0.299
        self.p_S = 0.1
        self.p_F = .08
        self.p_ln = []
        self.user_ids = user_ids
        try:
            self.users = np.unique(user_ids)
        except TypeError:
            self.users = self.unique(user_ids)
        self.answers = corrects
        self.handler = handler

    def unique(self, lst):
        new_lst = []
        for item in lst:
            print(item)
            if str(item) not in new_lst:
                new_lst.append(str(item))
        return np.array(new_lst)

    def set_ps(self, l, t, g, s, f, ol, ot, og, os):
        """ Setter for the precalculated parameters. """
        self.p_l0 = l
        self.p_T = t
        self.p_G = g
        self.p_S = s
        self.p_F = f

        # For old graphs
        self.p_ol0 = ol
        self.p_oT = ot
        self.p_oG = og
        self.p_oS = os


    def get_p_js(self, user_id, method='all', objective_id=None):
        """ Get and calculate the P(J_n).

        P(J_n) = P(J_l) - P(J_f). This is calculated with P(L), P(~L^T),
        P(~L^~T) and P(L^F).

        :param user_id: string; the id of the user for which we want P(J).
        :param method: string; How we handle multiple answers for one exercise.
        :param objective_id: string: For which objective_id we want answers.
        :returns: list of P(J) per answer.
        """
        # get the correct answers based on methods.
        user_answers, split = self.filter_answers(user_id, method,
                                                  objective_id)

        #calculate P(L_n) based on the answers.
        p_ln = self.calculate_ln(user_answers)
        # Calculate P(~l_n^T) and P(~L_n~T)
        p_not_ln_t = [(1 - ln) * self.p_T for ln in p_ln]
        p_not_ln_not_t = [(1 - ln) * (1 - self.p_T) for ln in p_ln]
        p_ln_f = [ln * self.p_F for ln in p_ln]
        p_ln_not_f = [ln * (1 - self.p_F) for ln in p_ln]

        p_jl = self.calculate_p_jl(user_answers, p_ln, p_not_ln_t,
                                   p_not_ln_not_t)
        p_jf = self.calculate_p_jf(user_answers, p_ln_f, p_not_ln_t,
                                   p_not_ln_not_t, p_ln_not_f)

        old_ln = self.calculate_ln(user_answers, old=True)
        old_not_ln_t = [(1 - ln) * self.p_oT for ln in old_ln]
        print(self.p_oT, old_not_ln_t)
        old_not_ln_not_t = [(1 - ln) * (1 - self.p_oT) for ln in old_ln]
        old_graph = self.calculate_p_jl(user_answers, old_ln, old_not_ln_t,
                                        old_not_ln_not_t, old=True)

        return [jl-jf for jl, jf in zip(p_jl, p_jf)], p_jl, p_jf, old_graph, \
               split, user_answers

    def filter_answers(self, user_id, method, objectives_id):
        """ Filter the answers on user, objective and method.

        :param user_id: string; the user ID.
        :param method: string; How we handle multiple answers on an exercise.
        :param objectives_id: string; The ID for the objectives.
        :return: list of answers to calculate P(J) for.
        """

        # Get the ID's of which data is to be selected
        self.chosen_ids = np.where(
            (self.handler.learn_obj_ids == objectives_id) &
            (self.user_ids == user_id))

        # Get the data.
        user_answers = self.answers[self.chosen_ids]
        user_objectives = self.handler.learn_obj_ids[self.chosen_ids]
        user_ids = self.user_ids[self.chosen_ids]
        user_excs = self.handler.exercise_ids[self.chosen_ids]
        user_dates = self.handler.dates[self.chosen_ids]
        user_abs = self.handler.ability_scores[self.chosen_ids]

        # Filter the data according to method.
        if method not in ['all', 'first', 'second', 'all but first', 'last']:
            raise NotImplementedError
        if method == 'first':
            user_answers = self.filter_all_but_first(user_answers, user_excs)
        if method == 'second':
            user_answers = self.filter_all_but_second(user_answers, user_excs)
        if method == 'all but first':
            user_answers = self.filter_first(user_answers, user_excs)
        if method == 'last':
            user_answers = self.filter_all_but_last(user_answers, user_excs)

        # Store the exercises.
        self.excs = user_excs
        self.dats = user_dates

        split = -1
        while user_abs[split+1] == 'NULL' and split<len(user_abs)-2:
            split = split+1
        split += 1
        if user_abs[split] == 'NULL':
            return user_answers, (None, '')
        return user_answers, (split, user_abs[split])

    def filter_all_but_first(self, answers, exercise_ids):
        """
        Filter out all but the first answers per exercise.

        :param answers:         list of answers, to be filtered.
        :param exercise_ids:    list of exercise ID's to indicate to which
            exercise answers belong.
        :return: list of all but the first answers for an exercise.
        """
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            # If it is the first time that I see this exercise:
            if exercise_ids[i] not in exercises_processed:
                # Keep answer.
                return_answers.append(answers[i])
                # Store exercise ID.
                exercises_processed.append(exercise_ids[i])
        return return_answers

    def filter_all_but_second(self, answers, exercise_ids):
        """
        Filter out all but the second answers per exercise.

        :param answers:         list of answers, to be filtered.
        :param exercise_ids:    list of exercise ID's to indicate to which
            exercise answers belong.
        :return: list of all but the second answers for an exercise.
        """
        exercises_processed = []
        exercises_processed_twice = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                if exercise_ids[i] not in exercises_processed_twice:
                    return_answers.append(answers[i])
                    exercises_processed_twice.append(exercise_ids[i])
        return return_answers

    def filter_first(self, answers, exercise_ids):
        """
        Filter out all the first answers per exercise.

        :param answers:         list of answers, to be filtered.
        :param exercise_ids:    list of exercise ID's to indicate to which
            exercise answers belong.
        :return: list of all the first answers for an exercise.
        """
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                exercises_processed.append(exercise_ids[i])
            else:
                return_answers.append(answers[i])
        return return_answers

    def filter_all_but_last(self, answers, exercise_ids):
        """
        Filter out all but the last answers per exercise.

        :param answers:         list of answers, to be filtered.
        :param exercise_ids:    list of exercise ID's to indicate to which
            exercise answers belong.
        :return: list of all but the last answers for an exercise.
        """
        # Reverse list, so last answers are first.
        answers = answers[::-1]
        exercises_processed = []
        return_answers = []
        for i in range(len(answers)):
            if exercise_ids[i] not in exercises_processed:
                return_answers.append(answers[i])
                exercises_processed.append(exercise_ids[i])
        return return_answers[::-1]

    def calculate_ln(self, answers, old=False):
        """
        Calculate P(L_n)
        :param answers: list of answers over which we calculate P(L_n)
        :return: a list of values of P(L_n)
        """
        p_ln = []
        for answer_id in range(len(answers)):
            if len(p_ln) == 0:
                k = self.p_l0
                if old is True:
                    k = self.p_ol0
            else:
                k = p_ln[-1]
            s = self.p_S
            g = self.p_G
            if old is True:
                s = self.p_oS
                g = self.p_oG
            if answers[answer_id] == 1:
                ln_prev_given_res = (k * (1 - s)) / (
                    (k * (1 - s)) + ((1 - k) * g))
            else:
                ln_prev_given_res = (k * s) / ((k * s) + ((1 - k) * (1 - g)))
            if old is False:
                p_ln.append(ln_prev_given_res * (1.0 - self.p_F) +
                            (1 - ln_prev_given_res) * self.p_T)
            else:
                p_ln.append(ln_prev_given_res +
                            (1 - ln_prev_given_res) * self.p_oT)

        return p_ln

    def calculate_p_jf(self, answers, ln_f, n_ln_t, n_ln_n_t, ln_nf):
        """
        Calculates P(J_f). Based upon the precalculated changes P(L_n^F),
        P(~L_n^T), P(~L_n^~T) and P(L_n^~F)

        :param answers:
        :param ln_f:
        :param n_ln_t:
        :param n_ln_n_t:
        :param ln_nf:
        :return:
        """
        p_jf = []
        for a_id in range(len(answers) - 2):
            p_ln_f = ln_f[a_id]
            g = self.p_G
            t = self.p_T
            s = self.p_S
            f = self.p_F
            if answers[a_id + 1] == 1:
                if answers[a_id + 2] == 1:                      # C C
                    a_l_nf =  (1-s)*(1-f)*(1-s) + (1-s)* f   * g
                    a_l_f =   g    * t   *(1-s) + g    *(1-t)* g
                    a_nl_t =  a_l_nf
                    a_nl_nt = a_l_f
                else:                                           # C ~C
                    a_l_nf =  (1-s)*(1-f)* s    + (1-s)* f   *(1-g)
                    a_l_f =   g    * t   * s    + g    *(1-t)*(1-g)
                    a_nl_t =  a_l_nf
                    a_nl_nt = a_l_f
            else:
                if answers[a_id + 2] == 1:                      # ~C C
                    a_l_nf =  s    *(1-f)*(1-s) + s    * f   * g
                    a_l_f =   (1-g)* t   *(1-s) + (1-g)*(1-t)* g
                    a_nl_t =  a_l_nf
                    a_nl_nt = a_l_f
                else:                                           # ~C ~C
                    a_l_nf =  s    *(1-f)* s    + s    * f   *(1-g)
                    a_l_f =   (1-g)* t   * s    + (1-g)*(1-t)*(1-g)
                    a_nl_t =  a_l_nf
                    a_nl_nt = a_l_f
            a12 = a_nl_t * n_ln_t[a_id] + a_nl_nt * n_ln_n_t[a_id] + \
                a_l_nf * ln_nf[a_id] + a_l_f * p_ln_f
            # a12 = a_l_nf * ln_nf[a_id] + a_l_f * p_ln_f
            p_jf.append(a_l_f * p_ln_f / a12)
        return p_jf

    def calculate_p_jl(self, answers, ln, n_ln_t, n_ln_n_t, old=False):
        """
        Calculate P(J_l). Was previous just P(J).

        :param answers:     Whether student answered correct.
        :param ln:          precalculated P(L_n)
        :param n_ln_t:      precalculated P(~L_n^T)
        :param n_ln_n_t:    precalculated P(~L_n^T)
        :return:            List of P(J_l) for every answer except the last
                            two.
        """
        p_jl = []
        for a_id in range(len(answers) - 2):
            p_l = ln[a_id]
            p_nl_t = n_ln_t[a_id]
            p_nl_nt = n_ln_n_t[a_id]
            if old is True:
                g = self.p_oG
                t = self.p_oT
                s = self.p_oS
            else:
                g = self.p_G
                t = self.p_T
                s = self.p_S
            if answers[a_id + 1] == 1:
                if answers[a_id + 2] == 1:  # RR
                    a_l = (1 - s) ** 2
                    a_nl_nt = g * (1 - t) * g + g * t * (1 - s)
                else:  # RW
                    a_l = s * (1 - s)
                    a_nl_nt = g * (1 - t) * (1 - g) + g * t * s
            else:
                if answers[a_id + 2] == 1:  # WR
                    a_l = s * (1 - s)
                    a_nl_nt = g * (1 - t) * (1 - g) + (1 - g) * t * (1 - s)
                else:  # WW
                    a_l = s ** 2
                    a_nl_nt = (1 - g) * (1 - t) * (1 - g) + (1 - g) * t * s
            a_nl_t = a_l
            a12 = p_l * a_l + p_nl_t * a_nl_t + p_nl_nt * a_nl_nt
            p_jl.append(a_nl_t * p_nl_t / a12)
        return p_jl

    def get_color_bars(self):
        # Gets the color corresponding to the boundaries. TODO: Update comment
        bounds = [0, 0, 0, 0, 0, 0, 0]
        colors = ['royalblue', 'darkorange', 'silver', 'gold', 'mediumblue',
                  'olivedrab']

        excs = self.excs
        dats = self.dats
        # find pre bound
        # print('finding pre-test')

        # print(e)
        try:
            e = excs[0]
            print(e)
            # if excs[1] in self.handler.pre_ids:
            while e in self.handler.pre_ids:
                bounds = [bounds[i] + 1 if i > 0 else bounds[i] for i in
                          range(len(bounds))]
                excs = excs[1:]
                dats = dats[1:]
                e = excs[0]

            d = dats[0]
            # Find class instruction
            # print('finding instruction exercises')
            while e in self.handler.c_in_ids:
                bounds = [bounds[i] + 1 if i > 1 else bounds[i] for i in
                          range(len(bounds))]
                excs = excs[1:]
                dats = dats[1:]
                e = excs[0]

            # Find class exercise
            # print('finding class exercises')
            while e in self.handler.c_ex_ids:
                bounds = [bounds[i] + 1 if i > 2 else bounds[i] for i in
                          range(len(bounds))]
                excs = excs[1:]
                d = dats[0]
                dats = dats[1:]
                e = excs[0]

            # Find class adaptive
            # print('finding class adaptive')
            while dats[0].strftime("%d") == d.strftime("%d") \
                    and not e == self.excs[0] \
                    and not (bounds[1] == bounds[2]
                             and bounds[2] == bounds[3]):
                bounds = [bounds[i] + 1 if i > 3 else bounds[i] for i in
                          range(len(bounds))]
                excs = excs[1:]
                dats = dats[1:]
                e = excs[0]

            # Find repetition adaptive and post test boundary
            # Done by finding all post-test exercises backwards.
            # print('finding repeated adaptive exercises')
            excs = excs[::-1]
            dats = dats[::-1]
            bounds[-1] = len(self.excs)
            bounds[-2] = bounds[-1]
            d = dats[0]
            found_post = False
            while dats[0].strftime("%d") == d.strftime("%d"):
                bounds[-2] -= 1
                if excs[0] in self.handler.post_ids:
                    found_post = True
                excs = excs[1:]
                dats = dats[1:]
            if found_post is False:
                bounds[-2] = bounds[-1]

        except IndexError:
            pass
        print(bounds)
        return bounds, colors


class ParameterExtractor:
    """
    Calculates the pre-calculated parameters.

    Has an option of brute force calculating the parameters or
    """
    def __init__(self):
        # params = [L0, T, G, S, F]
        self.params_min = [1e-15 for i in range(4)]
        self.params_max = [1.0, 1.0, 0.3, 0.1]

    def brute_force_params(self, answers, same, grain=100, L0_fix=None,
                           T_fix=None, G_fix=None, S_fix=None):
        """
        Check for every parameter what the best value is.

        if x_fix is None then the whole range will be tested.
        :param answers: the answers given by the students
        :param same: Whether the answers are switching to a new student.
        :param grain: integer defining the amount of values being checked
        :param L0_fix: integer; value of L0. if None, this will return best L0
        :param T_fix: integer; value of T. if None, this will return best T
        :param G_fix: integer; value of G. if None, this will return best G
        :param S_fix: integer; value of S. if None, this will return best S

        :return: values for L0, T, G and S that result in the lowest SSR.
        """
        # set ranges up
        best_l0 = L0_fix
        best_t = T_fix
        best_g = G_fix
        best_s = S_fix
        best_SSR = len(answers) * 999999999999991
        L0_range = self.get_range(L0_fix, 0, grain)
        T_range = self.get_range(T_fix, 1, grain)
        G_range = self.get_range(G_fix, 2, grain)
        S_range = self.get_range(S_fix, 3, grain)

        # Find best values
        for L0 in L0_range:
            # print('------------------------------------\nL0 is now at:{}'.format(L0))
            for T in T_range:
                for G in G_range:
                    for S in S_range:
                        # Get SSR for values
                        new_SSR = self.get_s_s_r(L0, T, G, S, answers,
                                                 same)

                        # check whether new value improves old values
                        if new_SSR <= best_SSR:
                            best_l0, best_t, best_g, best_s, \
                                best_SSR = [L0, T, G, S, new_SSR]
                            # print('best parameters now at L0:{}, ' +
                            #    'T:{}, G:{}, S:{}'.format(L0,
                            # 	 T, G, S))
        return best_l0, best_t, best_g, best_s

    def get_s_s_r(self, L0, T, G, S, answers, sames=None):
        """
        Calculate the Sum Squared Residu.

        This is a method that defines the fit of the parameters.
        :param L0: integer; value of L0
        :param T: integer; value of T
        :param G: integer; value of G
        :param S: integer; value of S
        :param answers: list of answers given by students
        :param sames: list of whether the answer is given by the same student.
        :return: float; Summed squared residu
        """
        SSR = 0.0
        S = max(1E-15, S)
        T = max(1E-15, T)
        G = max(1E-15, G)
        L0 = max(1E-15, L0)
        L = L0
        # Make sure that there is a list with sames.
        if sames is None:
            sames = np.ones(answers.size)
            sames[0] = 1

        # for every answer update the SSR.
        for same, answer in zip(sames, answers):
            if same == 0:  # New student so reset to initial chance of learning
                L = L0
            # print(L, T, G, S, F)
            SSR += (answer - (L * (1.0 - S) + (1.0 - L) * G)) ** 2
            if answer == 0:
                L_given_answer = (L * S) / ((L * S) + ((1.0 - L) * (1.0 - G)))
            else:
                L_given_answer = (L * (1.0 - S)) / (
                    (L * (1.0 - S)) + ((1.0 - L) * G))
            L = L_given_answer + (1.0 - L_given_answer) * T
        return SSR

    def get_range(self, possible_range, par_id, grain):
        """
        helperfunction to get the range for a parameter based on the grain.

        returns either a list of the whole possible values if that value is
        not set (possible_range=None) else it returns a list containing only
        once the value of the set value.
        :param possible_range: Either float with the preset value or None
        :param par_id: the id of the parameter to find the boundaries for it.
        :param grain: integer, how finegrained the range must be.
        :return:
        """
        if possible_range is None:
            return np.linspace(self.params_min[par_id],
                               self.params_max[par_id],
                               int(grain * self.params_max[par_id]),
                               endpoint=False)[1:]
        return [possible_range]

    def smart_ssr(self, answers, same, grain, iterations):
        best_l0 = self.brute_force_params(answers, same, grain,
                                          None, 0.0, 0.0, 0.0)[0]
        best_t = self.brute_force_params(answers, same, grain,
                                         0.0, None, 0.0, 0.0)[1]
        best_g = self.brute_force_params(answers, same, grain,
                                         0.0, 0.0, None, 0.0)[2]
        best_s = self.brute_force_params(answers, same, grain,
                                         0.0, 0.0, 0.0, None)[3]
        for i in range(iterations):
            print("best is {}".format([best_l0, best_t, best_g, best_s]))
            best_l0 = self.brute_force_params(answers, same, grain, None,
                                              best_t, best_g, best_s)[0]
            best_t = self.brute_force_params(answers, same, grain, best_l0,
                                             None, best_g, best_s)[1]
            best_g = self.brute_force_params(answers, same, grain, best_l0,
                                             best_t, None, best_s)[2]
            best_s = self.brute_force_params(answers, same, grain, best_l0,
                                             best_t, best_g, None)[3]
        return best_l0, best_t, best_g, best_s


class ParameterExtractor_with_forget:
    """
    Calculates the pre-calculated parameters.

    Has an option of brute force calculating the parameters or
    """
    def __init__(self):
        # params = [L0, T, G, S, F]
        self.params_min = [1e-15 for i in range(4)]
        self.params_min.append(0.01)
        self.params_max = [1.0, 1.0, 0.3, 0.1, 0.3]

    def brute_force_params(self, answers, same, grain=100, L0_fix=None,
                           T_fix=None, G_fix=None, S_fix=None, F_fix=None):
        """
        Check for every parameter what the best value is.

        if x_fix is None then the whole range will be tested.
        :param answers: the answers given by the students
        :param same: Whether the answers are switching to a new student.
        :param grain: integer defining the amount of values being checked
        :param L0_fix: integer; value of L0. if None, this will return best L0
        :param T_fix: integer; value of T. if None, this will return best T
        :param G_fix: integer; value of G. if None, this will return best G
        :param S_fix: integer; value of S. if None, this will return best S
        :param F_fix: integer; value of F. if None, this will return best F

        :return: values for L0, T, G, S and F that result in the lowest SSR.
        """
        # set ranges up
        best_l0 = L0_fix
        best_t = T_fix
        best_g = G_fix
        best_s = S_fix
        best_f = F_fix
        best_SSR = len(answers) * 999999999999991
        L0_range = self.get_range(L0_fix, 0, grain)
        T_range = self.get_range(T_fix, 1, grain)
        G_range = self.get_range(G_fix, 2, grain)
        S_range = self.get_range(S_fix, 3, grain)
        F_range = self.get_range(F_fix, 4, grain)

        # Find best values
        for L0 in L0_range:
            # print('------------------------------------\nL0 is now at:{}'.format(L0))
            for T in T_range:
                for G in G_range:
                    for S in S_range:
                        for F in F_range:
                            # Get SSR for values
                            new_SSR = self.get_s_s_r(L0, T, G, S, F, answers,
                                                     same)

                            # check whether new value improves old values
                            if new_SSR <= best_SSR:
                                best_l0, best_t, best_g, best_s, \
                                best_f, best_SSR = [L0, T, G, S, F, new_SSR]
                                # print('best parameters now at L0:{}, ' +
                                #    'T:{}, G:{}, S:{}'.format(L0,
                                # 	 T, G, S))
        return best_l0, best_t, best_g, best_s, best_f

    def get_s_s_r(self, L0, T, G, S, F_, answers, sames=None):
        """
        Calculate the Sum Squared Residu.

        This is a method that defines the fit of the parameters.
        :param L0: integer; value of L0
        :param T: integer; value of T
        :param G: integer; value of G
        :param S: integer; value of S
        :param F_: integer; value of F
        :param answers: list of answers given by students
        :param sames: list of whether the answer is given by the same student.
        :return: float; Summed squared residu
        """
        SSR = 0.0
        S = max(1E-15, S)
        T = max(1E-15, T)
        G = max(1E-15, G)
        L0 = max(1E-15, L0)
        F_ = max(1E-15, F_)
        L = L0
        # Make sure that there is a list with sames.
        if sames is None:
            sames = np.ones(answers.size)
            sames[0] = 1

        # for every answer update the SSR.
        for same, answer in zip(sames, answers):
            if same == 0:  # New student so reset to initial chance of learning
                L = L0
            # print(L, T, G, S, F)
            SSR += (answer - (L * (1.0 - S) + (1.0 - L) * G)) ** 2
            if answer == 0:
                L_given_answer = (L * S) / ((L * S) + ((1.0 - L) * (1.0 - G)))
            else:
                L_given_answer = (L * (1.0 - S)) / (
                    (L * (1.0 - S)) + ((1.0 - L) * G))
            L = L_given_answer * (1.0 - F_) + (1.0 - L_given_answer) * T
        return SSR

    def get_range(self, possible_range, par_id, grain):
        """
        helperfunction to get the range for a parameter based on the grain.

        returns either a list of the whole possible values if that value is
        not set (possible_range=None) else it returns a list containing only
        once the value of the set value.
        :param possible_range: Either float with the preset value or None
        :param par_id: the id of the parameter to find the boundaries for it.
        :param grain: integer, how finegrained the range must be.
        :return:
        """
        if possible_range is None:
            return np.linspace(self.params_min[par_id],
                               self.params_max[par_id],
                               int(grain * self.params_max[par_id]),
                               endpoint=False)[1:]
        return [possible_range]

    def smart_ssr_f(self, answers, same, grain, iterations):
        best_l0 = self.brute_force_params(answers, same, grain,
                                          None, 0.0, 0.0, 0.0, 0.0)[0]
        best_t = self.brute_force_params(answers, same, grain,
                                         0.0, None, 0.0, 0.0, 0.0)[1]
        best_g = self.brute_force_params(answers, same, grain,
                                         0.0, 0.0, None, 0.0, 0.0)[2]
        best_s = self.brute_force_params(answers, same, grain,
                                         0.0, 0.0, 0.0, None, 0.0)[3]
        best_f = self.brute_force_params(answers, same, grain,
                                         0.0, 0.0, 0.0, 0.0, None)[4]
        for i in range(iterations):
            print("best is {}".format([best_l0, best_t, best_g, best_s,
                                       best_f]))
            best_l0 = self.brute_force_params(answers, same, grain, None,
                                              best_t, best_g, best_s,
                                              best_f)[0]
            best_t = self.brute_force_params(answers, same, grain, best_l0,
                                             None, best_g, best_s, best_f)[1]
            best_g = self.brute_force_params(answers, same, grain, best_l0,
                                             best_t, None, best_s, best_f)[2]
            best_s = self.brute_force_params(answers, same, grain, best_l0,
                                             best_t, best_g, None, best_f)[3]
            best_f = self.brute_force_params(answers, same, grain, best_l0,
                                             best_t, best_g, best_s, None)[4]
        return best_l0, best_t, best_g, best_s, best_f


if __name__ == "__main__":  # TESTING
    ex = ParameterExtractor()
    dh = DataHandler()
    for loid in np.unique(dh.learn_obj_ids):
        print("getting parameters for goal {}".format(loid))
        dh.set_ps_correct_calc(loid)
        print("Best parameters are {}".format([dh.m2m.p_l0, dh.m2m.p_T,
                                               dh.m2m.p_G, dh.m2m.p_S,
                                               dh.m2m.p_F]))
